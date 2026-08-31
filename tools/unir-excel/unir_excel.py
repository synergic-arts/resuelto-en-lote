from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


def merge_excel(input_dir: Path, output_file: Path) -> int:
    files = sorted(path for path in input_dir.glob("*.xlsx") if path.resolve() != output_file.resolve())
    if not files:
        raise SystemExit("No se encontraron archivos .xlsx")

    output = Workbook()
    sheet = output.active
    sheet.title = "Combinado"
    header_written = False
    rows_written = 0

    for file in files:
        workbook = load_workbook(file, read_only=True, data_only=True)
        source = workbook.active
        rows = source.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        if not header_written:
            sheet.append(list(header) + ["archivo_origen"])
            header_written = True
        for row in rows:
            sheet.append(list(row) + [file.name])
            rows_written += 1
        workbook.close()

    output_file.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_file)
    return rows_written


def interactive() -> None:
    print("RESUELTO EN LOTE · UNIR EXCEL")
    input_value = input("Carpeta con los Excel: ").strip().strip('"')
    output_value = input("Archivo de salida [excel_combinado.xlsx]: ").strip().strip('"')
    output_value = output_value or "excel_combinado.xlsx"
    try:
        rows = merge_excel(Path(input_value), Path(output_value))
        print(f"\nLISTO: {rows} filas combinadas en {output_value}")
    except (Exception, SystemExit) as exc:
        print(f"\nERROR: {exc}")
    input("\nPulsa Enter para cerrar...")


def main() -> None:
    if len(sys.argv) == 1:
        interactive()
        return
    parser = argparse.ArgumentParser(description="Une hojas Excel con la misma cabecera")
    parser.add_argument("input_dir", type=Path)
    parser.add_argument("output_file", type=Path)
    args = parser.parse_args()
    rows = merge_excel(args.input_dir, args.output_file)
    print(f"OK: {rows} filas combinadas en {args.output_file}")


if __name__ == "__main__":
    main()
