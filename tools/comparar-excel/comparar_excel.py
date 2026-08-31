from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


def read_table(path: Path) -> tuple[list[str], dict[object, tuple[object, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(value) for value in next(rows)]
    except StopIteration as exc:
        raise SystemExit(f"El archivo está vacío: {path}") from exc
    records = {row[0]: tuple(row) for row in rows if row and row[0] is not None}
    workbook.close()
    return header, records


def compare_excel(left_path: Path, right_path: Path, output_path: Path) -> int:
    left_header, left = read_table(left_path)
    right_header, right = read_table(right_path)
    if left_header != right_header:
        raise SystemExit("Las cabeceras no coinciden")

    output = Workbook()
    sheet = output.active
    sheet.title = "Diferencias"
    sheet.append(["clave", "estado", "campo", "anterior", "nuevo"])
    differences = 0
    for key in sorted(set(left) | set(right), key=str):
        if key not in left:
            sheet.append([key, "AÑADIDO", "*", None, str(right[key])])
            differences += 1
        elif key not in right:
            sheet.append([key, "ELIMINADO", "*", str(left[key]), None])
            differences += 1
        else:
            for index, field in enumerate(left_header[1:], start=1):
                if left[key][index] != right[key][index]:
                    sheet.append([key, "CAMBIADO", field, left[key][index], right[key][index]])
                    differences += 1
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)
    return differences


def interactive() -> None:
    print("RESUELTO EN LOTE · COMPARAR EXCEL")
    original = input("Excel original: ").strip().strip('"')
    updated = input("Excel actualizado: ").strip().strip('"')
    output = input("Informe de salida [diferencias_excel.xlsx]: ").strip().strip('"')
    output = output or "diferencias_excel.xlsx"
    try:
        count = compare_excel(Path(original), Path(updated), Path(output))
        print(f"\nLISTO: {count} diferencias detectadas en {output}")
    except (Exception, SystemExit) as exc:
        print(f"\nERROR: {exc}")
    input("\nPulsa Enter para cerrar...")


def main() -> None:
    if len(sys.argv) == 1:
        interactive()
        return
    parser = argparse.ArgumentParser(description="Compara dos Excel usando la primera columna como clave")
    parser.add_argument("original", type=Path)
    parser.add_argument("nuevo", type=Path)
    parser.add_argument("salida", type=Path)
    args = parser.parse_args()
    count = compare_excel(args.original, args.nuevo, args.salida)
    print(f"OK: {count} diferencias guardadas en {args.salida}")


if __name__ == "__main__":
    main()
